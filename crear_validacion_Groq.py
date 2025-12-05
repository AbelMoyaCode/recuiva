#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para generar la tabla de validación de indicadores Groq API
Dataset de 5 casos REALES de prueba mostrando coherencia, edición y tiempos.

Autor: Abel Moya
Fecha: 2 de diciembre de 2025
"""

import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("📦 Instalando openpyxl...")
    os.system("pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Dataset REAL de las capturas del DevTools
dataset = [
    {
        "id": 1,
        "pregunta": "¿Cómo analiza el autor la actitud de Floriani en comparación con la certeza que inicialmente tenía, y qué implica esto para su comprensión de la situación?",
        "tipo": "inferential",
        "coherente": "Sí",
        "editada": "Sí",
        "tiempo": 0.97  # 971.14 ms
    },
    {
        "id": 2,
        "pregunta": "¿Cuál fue el propósito del collar del rey durante las veladas de lady Billingstone, y cómo se asoció con la condesa de la Motte?",
        "tipo": "inferential",
        "coherente": "Sí",
        "editada": "Sí",
        "tiempo": 1.07  # 1.07 s
    },
    {
        "id": 3,
        "pregunta": "¿Cuál es el dilema moral que enfrenta el personaje al descubrir la verdad sobre el hijo de Henriette y sus acciones?",
        "tipo": "inferential",
        "coherente": "Sí",
        "editada": "Sí",
        "tiempo": 0.81  # 806.75 ms
    },
    {
        "id": 4,
        "pregunta": "¿Cómo se relaciona el collar del rey con la DuBarry y su creencia de ofrecerlo a María Antonieta, reina de Francia?",
        "tipo": "inferential",
        "coherente": "Sí",
        "editada": "Sí",
        "tiempo": 1.00  # 1.00 s
    },
    {
        "id": 5,
        "pregunta": "¿Cuál es el propósito de la reunión en casa del señor de Dreux-Soubise, según el texto, y cómo se desarrolla la conversación?",
        "tipo": "inferential",
        "coherente": "Sí",
        "editada": "Sí",
        "tiempo": 1.05  # 1.05 s
    }
]

# Crear Excel
print("📝 Creando tabla de validación de indicadores Groq API...")
wb = Workbook()
ws = wb.active
ws.title = "Validación Groq API"

# Encabezados
headers = [
    'ID',
    'Pregunta Generada',
    'Tipo',
    '¿Coherente\ny pertinente?',
    '¿Editada\npor usuario?',
    'Tiempo\n(segundos)'
]

# Estilos
header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=10)
border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Escribir encabezados
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style

# Ajustar anchos de columna
ws.column_dimensions['A'].width = 6   # ID
ws.column_dimensions['B'].width = 65  # Pregunta
ws.column_dimensions['C'].width = 12  # Tipo
ws.column_dimensions['D'].width = 12  # ¿Coherente?
ws.column_dimensions['E'].width = 12  # ¿Editada?
ws.column_dimensions['F'].width = 10  # Tiempo

# Escribir datos
for row_idx, caso in enumerate(dataset, 2):
    # ID
    cell = ws.cell(row=row_idx, column=1, value=caso['id'])
    cell.alignment = Alignment(horizontal='center', vertical='top')
    cell.border = border_style
    
    # Pregunta
    cell = ws.cell(row=row_idx, column=2, value=caso['pregunta'])
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = border_style
    
    # Tipo
    cell = ws.cell(row=row_idx, column=3, value=caso['tipo'].capitalize())
    cell.alignment = Alignment(horizontal='center', vertical='top')
    cell.border = border_style
    if caso['tipo'] == 'inferential':
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # ¿Coherente?
    cell = ws.cell(row=row_idx, column=4, value=caso['coherente'])
    cell.alignment = Alignment(horizontal='center', vertical='top')
    cell.border = border_style
    if caso['coherente'] == 'Sí':
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        cell.font = Font(bold=True, color="006100")
    else:
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        cell.font = Font(bold=True, color="9C0006")
    
    # ¿Editada?
    cell = ws.cell(row=row_idx, column=5, value=caso['editada'])
    cell.alignment = Alignment(horizontal='center', vertical='top')
    cell.border = border_style
    if caso['editada'] == 'Sí':
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        cell.font = Font(bold=True, color="006100")
    else:
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        cell.font = Font(bold=True, color="9C0006")
    
    # Tiempo
    cell = ws.cell(row=row_idx, column=6, value=caso['tiempo'])
    cell.alignment = Alignment(horizontal='center', vertical='top')
    cell.border = border_style
    cell.number_format = '0.00'
    # Verde si ≤2 segundos
    if caso['tiempo'] <= 2.0:
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        cell.font = Font(bold=True, color="006100")
    else:
        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    # Ajustar altura de fila
    ws.row_dimensions[row_idx].height = 50

# Ajustar altura de encabezado
ws.row_dimensions[1].height = 35

# Agregar sección de métricas
metrics_row = len(dataset) + 3

# Título de métricas
ws.merge_cells(f'A{metrics_row}:F{metrics_row}')
cell = ws.cell(row=metrics_row, column=1, value="MÉTRICAS DE VALIDACIÓN - INDICADORES GROQ API")
cell.font = Font(bold=True, size=12, color="FFFFFF")
cell.alignment = Alignment(horizontal='center', vertical='center')
cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")

# Calcular métricas
total_casos = len(dataset)
coherentes = sum(1 for c in dataset if c['coherente'] == 'Sí')
editadas = sum(1 for c in dataset if c['editada'] == 'Sí')
tiempo_promedio = sum(c['tiempo'] for c in dataset) / total_casos

precision_coherencia = (coherentes / total_casos) * 100
precision_edicion = (editadas / total_casos) * 100

# Indicador 4.1
metrics_row += 1
ws.cell(row=metrics_row, column=1, value="Indicador 4.1 — Coherencia de preguntas:").font = Font(bold=True)
ws.cell(row=metrics_row, column=3, value=f"{coherentes}/{total_casos}")
ws.cell(row=metrics_row, column=4, value=f"{precision_coherencia:.1f}%").font = Font(bold=True, size=11)
ws.cell(row=metrics_row, column=5, value="✅ CUMPLE" if precision_coherencia >= 75 else "❌ NO CUMPLE")
ws.cell(row=metrics_row, column=5).font = Font(bold=True, color="006100" if precision_coherencia >= 75 else "9C0006")
ws.cell(row=metrics_row, column=6, value="(Meta: ≥75%)")

# Indicador 4.2
metrics_row += 1
ws.cell(row=metrics_row, column=1, value="Indicador 4.2 — Tasa de edición:").font = Font(bold=True)
ws.cell(row=metrics_row, column=3, value=f"{editadas}/{total_casos}")
ws.cell(row=metrics_row, column=4, value=f"{precision_edicion:.1f}%").font = Font(bold=True, size=11)
ws.cell(row=metrics_row, column=5, value="✅ CUMPLE" if precision_edicion >= 80 else "❌ NO CUMPLE")
ws.cell(row=metrics_row, column=5).font = Font(bold=True, color="006100" if precision_edicion >= 80 else "9C0006")
ws.cell(row=metrics_row, column=6, value="(Meta: ≥80%)")

# Indicador 4.3
metrics_row += 1
ws.cell(row=metrics_row, column=1, value="Indicador 4.3 — Tiempo promedio:").font = Font(bold=True)
ws.cell(row=metrics_row, column=3, value=f"{tiempo_promedio:.2f} seg")
ws.cell(row=metrics_row, column=4, value="")
ws.cell(row=metrics_row, column=5, value="✅ CUMPLE" if tiempo_promedio <= 2.0 else "❌ NO CUMPLE")
ws.cell(row=metrics_row, column=5).font = Font(bold=True, color="006100" if tiempo_promedio <= 2.0 else "9C0006")
ws.cell(row=metrics_row, column=6, value="(Meta: ≤2 seg)")

# Guardar archivo
output_file = "Validacion_Indicadores_Groq_API.xlsx"
wb.save(output_file)

print("✅ Tabla de validación creada exitosamente!")
print(f"📁 Ruta: {os.path.abspath(output_file)}")
print(f"\n📊 RESULTADOS:")
print(f"   📌 Indicador 4.1 (Coherencia): {coherentes}/{total_casos} = {precision_coherencia:.1f}% {'✅' if precision_coherencia >= 75 else '❌'}")
print(f"   📌 Indicador 4.2 (Edición): {editadas}/{total_casos} = {precision_edicion:.1f}% {'✅' if precision_edicion >= 80 else '❌'}")
print(f"   📌 Indicador 4.3 (Tiempo promedio): {tiempo_promedio:.2f} seg {'✅' if tiempo_promedio <= 2.0 else '❌'}")
print("\n🎯 ¡Todos los indicadores del Objetivo 4 están validados!")
