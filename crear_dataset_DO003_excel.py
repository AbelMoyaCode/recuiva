#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para generar el dataset DO-003 en formato Excel (.xlsx)
con formato profesional y compatible con tildes/ñ.

Autor: Abel Moya
Fecha: 2 de diciembre de 2025
"""

import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl no está instalado.")
    print("📦 Instalando openpyxl...")
    os.system("pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Datos del dataset DO-003
dataset = [
    (1, "¿Qué era el famoso collar de la reina en la historia?", 
     "Era un collar de diamantes muy costoso encargado originalmente para María Antonieta.", 
     "Correcto"),
    (2, "¿Qué papel tuvo el cardenal de Rohan en la estafa?", 
     "Era un cardenal manipulado para creer que la reina quería que él comprara el collar.", 
     "Correcto"),
    (3, "¿Por qué la reputación de María Antonieta ya estaba dañada antes del escándalo del collar?", 
     "Porque circulaban muchos chismes sobre su vida lujosa y alejamiento del pueblo.", 
     "Parcial"),
    (4, "¿Quién era Jeanne de Valois-Saint-Rémy y cómo participó en el engaño?", 
     "Era una falsa condesa que planeó la estafa y fingió ser cercana a la reina.", 
     "Correcto"),
    (5, "¿Qué engaño se realizó en los jardines de Versalles durante la noche?", 
     "Hicieron que el cardenal creyera reunirse con una mujer disfrazada de María Antonieta.", 
     "Parcial"),
    (6, "¿Por qué los joyeros confiaron en que recibirían el pago del collar?", 
     "Porque había un contrato donde parecía que la reina aprobaba la compra.", 
     "Parcial"),
    (7, "¿Qué hizo Jeanne con el collar una vez que lo obtuvo?", 
     "Mandó desarmar el collar y vender los diamantes.", 
     "Correcto"),
    (8, "¿Cómo se enteró finalmente María Antonieta de lo que estaba ocurriendo?", 
     "Los joyeros acudieron a reclamar el pago y estalló el escándalo.", 
     "Correcto"),
    (9, "¿Qué consecuencias políticas tuvo el caso del collar?", 
     "Aumentó el odio hacia la monarquía y debilitó la imagen pública de la reina.", 
     "Parcial"),
    (10, "¿Por qué se considera que el caso del collar influyó en la Revolución Francesa?", 
     "Reforzó la idea de que la corte era corrupta y ajena al pueblo.", 
     "Parcial"),
    (11, "¿Qué buscaba el cardenal de Rohan al participar en el escándalo del collar?", 
     "Reconquistar el favor político de la reina.", 
     "Parcial"),
    (12, "¿Cómo se usaron las cartas falsas en el plan?", 
     "Eran cartas falsificadas para que el cardenal creyera que hablaba con la reina.", 
     "Parcial"),
    (13, "¿Qué papel tuvo la mujer que se hizo pasar por la reina en el jardín?", 
     "Solo apareció un momento para engañar al cardenal.", 
     "Parcial"),
    (14, "¿Por qué los joyeros aceptaron rebajar el precio del collar?", 
     "Jeanne los convenció de que sería más fácil que la reina pagara luego.", 
     "Parcial"),
    (15, "¿Qué pasó con Jeanne después de que estalló el escándalo?", 
     "Fue juzgada públicamente y terminó perdiendo credibilidad.", 
     "Parcial"),
    (16, "¿Qué representaba el collar para la sociedad francesa?", 
     "Era un símbolo del exceso y lujo de la nobleza.", 
     "Parcial"),
    (17, "¿Qué muestra el caso del collar sobre la relación entre la corte y el pueblo?", 
     "Que había mucha desconfianza hacia la realeza.", 
     "Parcial"),
    (18, "¿Qué oportunidad aprovechó la prensa para criticar a la reina?", 
     "El escándalo del collar se usó para atacar su imagen.", 
     "Parcial"),
    (19, "¿Cómo el cardenal de Rohan organizó el engaño a la reina?", 
     "Porque Jeanne lo manipuló haciéndole creer que la reina lo quería de vuelta.", 
     "Incorrecto"),
    (20, "¿Qué ocurrió al final con el collar cuando la reina quiso quemarlo?", 
     "La reina mandó destruirlo para demostrar inocencia.", 
     "Incorrecto"),
]

# Crear libro de Excel
print("📝 Creando archivo Excel del dataset DO-003...")
wb = Workbook()
ws = wb.active
ws.title = "DO-003 Collar Reina"

# Configurar encabezados
headers = ['ID', 'Pregunta', 'Respuesta del usuario', 'Score %', 
            'Clasificación Auto', 'Clasificación Manual', '¿Coincide?']

# Estilos
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Escribir encabezados
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_style

# Ajustar anchos de columna
ws.column_dimensions['A'].width = 6   # ID
ws.column_dimensions['B'].width = 50  # Pregunta
ws.column_dimensions['C'].width = 50  # Respuesta
ws.column_dimensions['D'].width = 10  # Score %
ws.column_dimensions['E'].width = 18  # Clasificación Auto
ws.column_dimensions['F'].width = 18  # Clasificación Manual
ws.column_dimensions['G'].width = 12  # ¿Coincide?

# Escribir datos
for row_idx, (id_caso, pregunta, respuesta, clasificacion_manual) in enumerate(dataset, 2):
    # ID
    cell = ws.cell(row=row_idx, column=1, value=id_caso)
    cell.alignment = Alignment(horizontal='center', vertical='top')
    cell.border = border_style
    
    # Pregunta
    cell = ws.cell(row=row_idx, column=2, value=pregunta)
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = border_style
    
    # Respuesta
    cell = ws.cell(row=row_idx, column=3, value=respuesta)
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = border_style
    
    # Score % (vacío para llenar)
    cell = ws.cell(row=row_idx, column=4, value="")
    cell.alignment = Alignment(horizontal='center', vertical='top')
    cell.border = border_style
    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Clasificación Auto (vacío para llenar)
    cell = ws.cell(row=row_idx, column=5, value="")
    cell.alignment = Alignment(horizontal='center', vertical='top')
    cell.border = border_style
    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Clasificación Manual
    cell = ws.cell(row=row_idx, column=6, value=clasificacion_manual)
    cell.alignment = Alignment(horizontal='center', vertical='top')
    cell.border = border_style
    
    # ¿Coincide? (vacío para llenar)
    cell = ws.cell(row=row_idx, column=7, value="")
    cell.alignment = Alignment(horizontal='center', vertical='top')
    cell.border = border_style
    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # Ajustar altura de fila
    ws.row_dimensions[row_idx].height = 40

# Ajustar altura de encabezado
ws.row_dimensions[1].height = 35

# Agregar filas de métricas al final
metrics_row = len(dataset) + 3
ws.merge_cells(f'A{metrics_row}:C{metrics_row}')
cell = ws.cell(row=metrics_row, column=1, value="MÉTRICAS DE VALIDACIÓN")
cell.font = Font(bold=True, size=12)
cell.alignment = Alignment(horizontal='center', vertical='center')
cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

# Fila para total de coincidencias
metrics_row += 1
ws.cell(row=metrics_row, column=1, value="Total Coincidencias:")
ws.cell(row=metrics_row, column=2, value="=COUNTIF(G2:G21,\"Sí\")")
ws.cell(row=metrics_row, column=3, value="/ 20")

# Fila para precisión general
metrics_row += 1
ws.cell(row=metrics_row, column=1, value="Precisión General:")
ws.cell(row=metrics_row, column=2, value="=COUNTIF(G2:G21,\"Sí\")/20")
ws.cell(row=metrics_row, column=2).number_format = '0.00%'
ws.cell(row=metrics_row, column=3, value="(Meta: ≥75%)")

# Fila para correctas/parciales
metrics_row += 1
ws.cell(row=metrics_row, column=1, value="Correctas/Parciales correctas:")
ws.cell(row=metrics_row, column=2, value="(Calcular manualmente)")
ws.cell(row=metrics_row, column=3, value="(Meta: ≥70%)")

# Guardar archivo
output_file = "DO-003_collar_reina.xlsx"
wb.save(output_file)

print("✅ Archivo Excel creado exitosamente!")
print(f"📁 Ruta: {os.path.abspath(output_file)}")
print(f"📊 Total de casos: {len(dataset)}")
print(f"📍 Distribución:")
print(f"   - Correctos: {sum(1 for r in dataset if r[3] == 'Correcto')}")
print(f"   - Parciales: {sum(1 for r in dataset if r[3] == 'Parcial')}")
print(f"   - Incorrectos: {sum(1 for r in dataset if r[3] == 'Incorrecto')}")
print("")
print("🔹 Las celdas amarillas son para que las llenes con:")
print("   - Score % (del sistema)")
print("   - Clasificación Auto (del sistema)")
print("   - ¿Coincide? (Sí/No comparando columnas E y F)")
print("")
print("📈 Las métricas se calcularán automáticamente al final.")
